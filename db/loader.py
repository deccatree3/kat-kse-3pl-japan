"""
월별 raw 데이터를 파싱하여 DB에 적재하는 로더.
사용법: python db/loader.py 202604
"""
import sys
import os
import glob
import sqlite3
import openpyxl

DB_PATH = os.path.join(os.path.dirname(__file__), "logistics.db")

# 청구내역 확인서의 항목 순서 및 카테고리 매핑
CATEGORY_MAP = {
    1: ('inbound_plt', '입고,재고등록(PLT)'),
    2: ('inbound_ctn', '입고,재고등록(CTN)'),
    3: ('inbound_pcs', '입고,재고등록(PCS)'),
    4: ('storage', '보관료'),
    5: ('picking_pcs', '피킹요금_PCS'),
    6: ('picking_ctn', '피킹요금_CTN'),
    7: ('b2c_handling', 'B2C 출하작업수수료'),
    8: ('b2b_handling', 'B2B 출하작업수수료'),
    9: ('cushion', '완충자재/Material Surcharge'),
    10: ('set_work', '세트 작업 비용'),
    11: ('labeling', '라벨링 작업 비용'),
    12: ('repalletize', 'Repalletizing 비용'),
    13: ('truck_load', '출하작업비용_PLT'),
    14: ('box_60', '40~60size 포장박스'),
    15: ('box_80', '80size 포장박스'),
    16: ('box_120', '100~120size 포장박스'),
    17: ('box_140', '140size 포장박스'),
    18: ('ship_60', 'B2C 배송료(~60사이즈)'),
    19: ('ship_80', 'B2C 배송료(80사이즈)'),
    20: ('ship_100', 'B2C 배송료(100사이즈)'),
    21: ('ship_120_140', 'B2C 배송료(120,140사이즈)'),
    22: ('ship_160', 'B2C 배송료(160사이즈)'),
    23: ('okinawa_relay', '오키나와 및 낙도 중계료'),
    24: ('ship_60_oki', 'B2C 배송료(~60사이즈)_沖縄'),
    25: ('ship_80_oki', 'B2C 배송료(80사이즈)_沖縄'),
    26: ('ship_100_oki', 'B2C 배송료(100사이즈)_沖縄'),
}


def parse_xlsx(year_month: str) -> dict:
    """Excel 상세 파일에서 검증용 데이터 추출"""
    raw_dir = f"raw/{year_month}"
    xlsx_files = glob.glob(os.path.join(raw_dir, "*.xlsx"))
    if not xlsx_files:
        print(f"No xlsx file found in {raw_dir}")
        return {}

    wb = openpyxl.load_workbook(xlsx_files[0], data_only=True)
    result = {}

    # 入庫 시트
    if '入庫' in wb.sheetnames:
        ws = wb['入庫']
        ctn_total, plt_total, pcs_total = 0, 0, 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, values_only=False):
            b = row[1].value  # PCS
            c = row[2].value  # CTN
            d = row[3].value  # PLT
            if isinstance(b, (int, float)):
                pcs_total += int(b)
            if isinstance(c, (int, float)):
                ctn_total += int(c)
            if isinstance(d, (int, float)):
                plt_total += int(d)
        result['inbound_ctn'] = ctn_total
        result['inbound_plt'] = plt_total
        result['inbound_pcs'] = pcs_total

    # B2C出荷集計 시트
    if 'B2C出荷集計' in wb.sheetnames:
        ws = wb['B2C出荷集計']
        sizes = {}
        total_orders = 0
        total_qty = 0
        relay_total = 0
        relay_count = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            a_val = row[0].value
            c_val = row[2].value
            e_val = row[4].value
            f_val = row[5].value
            if a_val and isinstance(a_val, (int, float)) and row[1].value:
                total_orders += 1
                if isinstance(c_val, (int, float)):
                    total_qty += int(c_val)
                if isinstance(e_val, (int, float)):
                    sizes[int(e_val)] = sizes.get(int(e_val), 0) + 1
                if isinstance(f_val, (int, float)) and f_val > 0:
                    relay_total += int(f_val)
                    relay_count += 1
        result['b2c_orders'] = total_orders
        result['total_picking_pcs'] = total_qty
        result['size_breakdown'] = sizes
        result['okinawa_relay_fee'] = relay_total
        result['okinawa_count'] = relay_count

    # B2B出荷 시트
    if 'B2B出荷' in wb.sheetnames:
        ws = wb['B2B出荷']
        b2b_count = 0
        b2b_ctn = 0
        b2b_labels = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            if row[0].value and isinstance(row[0].value, (int, float)):
                b2b_count += 1
                if isinstance(row[2].value, (int, float)):
                    b2b_ctn += int(row[2].value)
                if isinstance(row[5].value, (int, float)):
                    b2b_labels += int(row[5].value)
        result['b2b_orders'] = b2b_count
        result['b2b_ctn'] = b2b_ctn
        result['b2b_labels'] = b2b_labels

    wb.close()
    return result


def print_verification(year_month: str):
    """raw 데이터를 파싱하여 검증 결과 출력 (DB 적재 전 확인용)"""
    import sys, io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    data = parse_xlsx(year_month)
    if not data:
        return

    print(f"\n=== {year_month} 검증 데이터 ===")
    print(f"입고 CTN: {data.get('inbound_ctn', 0)}")
    print(f"입고 PLT: {data.get('inbound_plt', 0)}")
    print(f"B2C 주문 수: {data.get('b2c_orders', 0)}")
    print(f"피킹 PCS: {data.get('total_picking_pcs', 0)}")
    print(f"사이즈별: {data.get('size_breakdown', {})}")
    print(f"오키나와 중계료: JPY {data.get('okinawa_relay_fee', 0)} ({data.get('okinawa_count', 0)}건)")
    print(f"B2B 주문: {data.get('b2b_orders', 0)} ({data.get('b2b_ctn', 0)} CTN)")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python db/loader.py <YYYYMM>")
        print("Example: python db/loader.py 202604")
        sys.exit(1)

    ym = sys.argv[1]
    print_verification(ym)
