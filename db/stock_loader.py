"""
재고/출고 raw 파일을 DB에 적재.
사용: python db/stock_loader.py
"""
import os
import sys
import io
import json
import glob
import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

# 모듈 경로 (patched: package import)
_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
if _THIS_DIR not in sys.path:
    sys.path.insert(0, _THIS_DIR)
import pg as _pg  # db/pg.py

BASE_DIR = os.path.dirname(_THIS_DIR)
APP_CFG_PATH = os.path.join(BASE_DIR, "config.json")


def _load_raw_dir():
    default = os.path.join(BASE_DIR, "raw")
    if not os.path.exists(APP_CFG_PATH):
        return default
    try:
        with open(APP_CFG_PATH, "r", encoding="utf-8") as f:
            return json.load(f).get("raw_dir", default)
    except Exception:
        return default


RAW_DIR = _load_raw_dir()

SCHEMA = """
CREATE TABLE IF NOT EXISTS shipments (
    waybill TEXT,
    sku_code TEXT,
    order_no TEXT,
    ship_date TEXT,
    sku_name TEXT,
    qty INTEGER,
    ship_type TEXT,
    source_file TEXT,
    loaded_at TEXT,
    PRIMARY KEY (waybill, sku_code)
);

CREATE TABLE IF NOT EXISTS stock_snapshots (
    snapshot_date TEXT,
    sku_code TEXT,
    sku_name TEXT,
    total_qty INTEGER,
    available_qty INTEGER,
    source_file TEXT,
    loaded_at TEXT,
    PRIMARY KEY (snapshot_date, sku_code)
);

CREATE TABLE IF NOT EXISTS stock_load_meta (
    key TEXT PRIMARY KEY,
    value TEXT
);

CREATE INDEX IF NOT EXISTS idx_shipments_sku ON shipments(sku_code);
CREATE INDEX IF NOT EXISTS idx_shipments_date ON shipments(ship_date);
"""


def ensure_schema(conn):
    with conn.cursor() as cur:
        for stmt in SCHEMA.split(';'):
            stmt = stmt.strip()
            if stmt:
                cur.execute(stmt)
    conn.commit()


def load_order_file(path, conn):
    """ORDER_LIST_*.xlsx 파일 1개 → shipments 테이블"""
    ship_type = 'B2B' if 'B2B' in os.path.basename(path).upper() else 'B2C'
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    rows = []
    now = datetime.datetime.now().isoformat(timespec='seconds')
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or row[1] is None:
            continue
        order_no = row[1]
        waybill = row[35]
        ship_date = row[4] or row[2]
        sku_code = row[44]
        sku_name = row[45]
        qty = row[52]
        if not sku_code or not qty:
            continue
        if not waybill:
            waybill = f"NOWB_{order_no}"
        rows.append((
            str(waybill), str(sku_code), str(order_no),
            str(ship_date)[:8] if ship_date else None,
            sku_name, int(qty), ship_type,
            os.path.basename(path), now,
        ))
    wb.close()

    with conn.cursor() as cur:
        cur.executemany("""
            INSERT INTO shipments
            (waybill, sku_code, order_no, ship_date, sku_name, qty, ship_type, source_file, loaded_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (waybill, sku_code) DO UPDATE SET
                order_no = EXCLUDED.order_no,
                ship_date = EXCLUDED.ship_date,
                sku_name = EXCLUDED.sku_name,
                qty = EXCLUDED.qty,
                ship_type = EXCLUDED.ship_type,
                source_file = EXCLUDED.source_file,
                loaded_at = EXCLUDED.loaded_at
        """, rows)
    return len(rows)


def load_stock_file(path, conn):
    """재고현황_*.xlsx 1개 → stock_snapshots (snapshot_date = 파일명에서 추출)"""
    base = os.path.basename(path)
    datestr = base.replace('재고현황 내역_', '').replace('.xlsx', '')[:6]
    try:
        yy = int(datestr[:2])
        mm = int(datestr[2:4])
        dd = int(datestr[4:6])
        snapshot_date = f"20{yy:02d}-{mm:02d}-{dd:02d}"
    except ValueError:
        snapshot_date = datetime.date.today().isoformat()

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    now = datetime.datetime.now().isoformat(timespec='seconds')
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        sku_code = row[9]
        if not sku_code:
            continue
        rows.append((
            snapshot_date, str(sku_code), row[10],
            int(row[11] or 0), int(row[12] or 0),
            base, now,
        ))
    wb.close()

    with conn.cursor() as cur:
        cur.executemany("""
            INSERT INTO stock_snapshots
            (snapshot_date, sku_code, sku_name, total_qty, available_qty, source_file, loaded_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (snapshot_date, sku_code) DO UPDATE SET
                sku_name = EXCLUDED.sku_name,
                total_qty = EXCLUDED.total_qty,
                available_qty = EXCLUDED.available_qty,
                source_file = EXCLUDED.source_file,
                loaded_at = EXCLUDED.loaded_at
        """, rows)
    return snapshot_date, len(rows)


def rebuild_all():
    raw_dir = _load_raw_dir()
    conn = _pg.connect()
    ensure_schema(conn)

    order_files = []
    for pat in ["ORDER_LIST_*.xlsx", os.path.join("*", "ORDER_LIST_*.xlsx")]:
        order_files.extend(glob.glob(os.path.join(raw_dir, pat)))

    total_rows = 0
    for f in order_files:
        n = load_order_file(f, conn)
        print(f"[ORDER] {os.path.basename(f)}: {n} rows")
        total_rows += n

    stock_files = glob.glob(os.path.join(raw_dir, "재고현황*.xlsx"))
    latest_snapshot = None
    for f in stock_files:
        snapshot_date, n = load_stock_file(f, conn)
        print(f"[STOCK] {os.path.basename(f)}: {n} rows, snapshot={snapshot_date}")
        if latest_snapshot is None or snapshot_date > latest_snapshot:
            latest_snapshot = snapshot_date

    now = datetime.datetime.now().isoformat(timespec='seconds')
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO stock_load_meta (key, value) VALUES ('last_loaded_at', %s)
            ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value
        """, (now,))
        if latest_snapshot:
            cur.execute("""
                INSERT INTO stock_load_meta (key, value) VALUES ('latest_snapshot', %s)
                ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value
            """, (latest_snapshot,))

    conn.commit()
    conn.close()

    return {
        'order_files': len(order_files),
        'shipment_rows': total_rows,
        'stock_files': len(stock_files),
        'latest_snapshot': latest_snapshot,
        'loaded_at': now,
    }


if __name__ == "__main__":
    result = rebuild_all()
    print(f"\n완료: {result}")
