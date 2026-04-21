"""
QSM detail.csv → KSE OMS Outbound 양식 변환.
QSM brief.csv + 송장번호 매핑 → QSM 업로드용 CSV 생성.
"""
import csv
import copy
import io
import os
import sys
import datetime
from typing import List, Dict, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

_THIS = os.path.dirname(os.path.abspath(__file__))
OUTBOUND_TEMPLATE = os.path.join(_THIS, "templates", "outbound_template.xlsx")

# db/pg.py import
_THIS = os.path.dirname(os.path.abspath(__file__))
_BASE = os.path.dirname(_THIS)
if os.path.join(_BASE, "db") not in sys.path:
    sys.path.insert(0, os.path.join(_BASE, "db"))
import pg

# Outbound 51컬럼 순서 (헤더는 2줄 합침: 日本語\n영문코드)
OUTBOUND_HEADERS = [
    ("倉庫コード", "CTKEY"),
    ("荷主コード", "OWKEY"),
    ("出庫予定日", "OR_HDDATE"),
    ("注文日", "OR_DATE"),
    ("注文タイプ", "ORHDTYPE"),
    ("商品コード", "ICMPKEY"),
    ("商品オプション名称", "IC_OPTION"),
    ("商品単位コード", "ICUTKEY"),
    ("代替コード", "SBKEY"),
    ("物流グループコード", "LOGGRPCD"),
    ("販売先コード", "STORE_KEY"),
    ("単位", "UOM"),
    ("予定数量", "EXQTY"),
    ("生産日", "PRODUCTDATE"),
    ("有効日", "EXPIREDATE"),
    ("注文番号", "EXTERNORDERKEY"),
    ("仕入先コード", "ACKEY"),
    ("仕入先名/受取人名", "ACNAME"),
    ("電話番号", "TEL"),
    ("携帯電話番号", "CP"),
    ("FAX番号", "FAX"),
    ("担当者", "CONTACT"),
    ("国コード", "COUNTRYCODE"),
    ("郵便番号コード", "ZCKEY"),
    ("基本住所", "ADDRESS1"),
    ("詳細住所", "ADDRESS2"),
    ("都市", "CITY"),
    ("都道府県(州)", "STATE"),
    ("配送会社", "DLCOMPANY"),
    ("注文配送運賃タイプ", "ODPAYTYPE"),
    ("配達指定日", "DLDATE"),
    ("配達時間帯", "DLTIME"),
    ("注文担当者", "OR_USER_ID"),
    ("注文先名", "ORNAME"),
    ("注文先電話番号", "ORTEL"),
    ("注文先FAX番号", "ORFAX"),
    ("注文先担当者", "ORCONTACT"),
    ("注文先国コード", "ORCOUNTRYCODE"),
    ("注文先郵便番号", "ORZCKEY"),
    ("注文先基本住所", "ORADDRESS1"),
    ("注文先詳細住所", "ORADDRESS2"),
    ("注文先都市", "ORCITY"),
    ("注文先都道府県(州)", "ORSTATE"),
    ("単位原価", "COSTPRICE"),
    ("販売価格", "SALEPRICE"),
    ("ベンダーコード", "VDKEY"),
    ("集合梱包情報", "PACKAGESOURCE"),
    ("コメント1", "COMMENTS1"),
    ("コメント2", "COMMENTS2"),
    ("TC/DC", "ATTRIBUTE1"),
    ("一般/保税", "ATTRIBUTE2"),
]


def load_mappings() -> Dict[Tuple[str, str], Dict]:
    """DB에서 상품 매핑 로드. key=(상품명, 옵션)"""
    conn = pg.connect(autocommit=True)
    with conn.cursor() as cur:
        cur.execute("""
            SELECT qoo10_name, qoo10_option, item_codes, sku_codes, quantities, enabled
            FROM qoo10_product_mapping
        """)
        rows = cur.fetchall()
    conn.close()
    result = {}
    for r in rows:
        result[(r[0], r[1] or '')] = {
            'item_codes': (r[2] or '').split(','),
            'sku_codes': (r[3] or '').split(','),
            'quantities': [int(x) for x in (r[4] or '1').split(',')],
            'enabled': bool(r[5]),
        }
    return result


def parse_qsm_csv(content: bytes) -> List[Dict]:
    """QSM detail.csv bytes → list of dict"""
    text = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def normalize_postal(code: str) -> str:
    """QSM은 '289-1733 형식 → 289-1733"""
    if not code:
        return ''
    return code.lstrip("'").strip()


def clean_special_chars(text: str) -> str:
    """
    템플릿 VBA Module1.CleanSpecialChars 포팅.

    삭제 범위:
      - U+2000~U+206F (8192~8303): 일반 문장 부호 (특수 공백, em/en dash, curly quotes 등)
      - U+2600~U+26FF (9728~9983): 기호 및 도형 (★ ◆ ♠ ✓ 등)

    보존: 그 외 모든 문자 (일본어/한자/숫자/하이픈-/전각장음ー/ｰ 등).
    """
    if not text:
        return ''
    out_chars = []
    for ch in text:
        cp = ord(ch)
        if 8192 <= cp <= 8303 or 9728 <= cp <= 9983:
            continue  # 삭제
        out_chars.append(ch)
    return ''.join(out_chars)


def normalize_order_date(qsm_date: str) -> str:
    """2026/04/15 19:12:16 → 20260415"""
    if not qsm_date:
        return ''
    try:
        dt = datetime.datetime.strptime(qsm_date.strip(), '%Y/%m/%d %H:%M:%S')
        return dt.strftime('%Y%m%d')
    except ValueError:
        # 이미 YYYYMMDD일 수 있음
        digits = ''.join(c for c in qsm_date if c.isdigit())
        return digits[:8] if len(digits) >= 8 else qsm_date


def generate_outbound_rows(qsm_rows: List[Dict], mappings: Dict) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    """
    QSM detail 행들 → Outbound 행들 변환.
    Power Query 로직 준수:
      1. 취합대상(enabled)=y만 필터
      2. 품목코드("SKU1,qty1,SKU2,qty2") split & expand → N SKU 행
      3. 予定数量 = QSM수량 × 매핑 SKU당수량
      4. 정렬: 장바구니번호 ASC, 주문번호 ASC, 품목(SKU) ASC
      5. 注文番号는 장바구니번호 사용 (같은 장바구니 = 합포장)
      6. 주소는 VBA CleanSpecialChars 적용 (★ ◆ 등 제거)
    반환: (출고 행들, 미매핑/에러 행들, 주소 정제 변경 이력)
    """
    today = datetime.date.today().strftime('%Y%m%d')
    outbound_rows = []
    errors = []
    addr_changes = []  # 정제로 변경된 주소 목록 (사용자 확인용)

    for q in qsm_rows:
        name = (q.get('상품명') or '').strip()
        option = (q.get('옵션정보') or '').strip()
        qsm_qty = int(q.get('수량', '1') or 1)
        cart_no = (q.get('장바구니번호') or '').strip()
        order_no = (q.get('주문번호') or '').strip()

        m = mappings.get((name, option))
        if m is None:
            errors.append({
                '장바구니번호': cart_no, '주문번호': order_no,
                '상품명': name, '옵션정보': option,
                '원인': '상품 매핑 없음',
            })
            continue
        if not m['enabled']:
            errors.append({
                '장바구니번호': cart_no, '주문번호': order_no,
                '상품명': name, '옵션정보': option,
                '원인': '매핑 비활성(취급 안함)',
            })
            continue

        # 주소 특수문자 정제 (VBA CleanSpecialChars)
        orig_addr = (q.get('주소') or '').strip()
        clean_addr = clean_special_chars(orig_addr)
        if orig_addr != clean_addr:
            addr_changes.append({
                '장바구니번호': cart_no, '주문번호': order_no,
                '원본주소': orig_addr, '정제주소': clean_addr,
            })

        # 전화번호: 수취인핸드폰 > 수취인전화 (값이 "-"면 skip)
        tel_cands = [
            q.get('수취인핸드폰번호', '').strip(),
            q.get('수취인전화번호', '').strip(),
        ]
        tel = next((t for t in tel_cands if t and t != '-'), '')

        # SKU별 1행 생성 (세트 상품은 N행으로 분할)
        for sku_code, sku_unit_qty in zip(m['sku_codes'], m['quantities']):
            if not sku_code or sku_code == '-':
                continue
            row = {h[0]: '' for h in OUTBOUND_HEADERS}
            row['倉庫コード'] = 'KE00003'
            row['荷主コード'] = 'katchers'
            row['出庫予定日'] = today
            row['注文日'] = normalize_order_date(q.get('주문일', ''))
            row['商品コード'] = sku_code
            row['予定数量'] = sku_unit_qty * qsm_qty  # 핵심: 매핑수량 × QSM수량
            row['注文番号'] = cart_no  # 장바구니번호 사용 (합포장)
            row['仕入先名/受取人名'] = q.get('수취인명', '')
            row['電話番号'] = tel
            row['国コード'] = 'JPN'
            row['郵便番号コード'] = normalize_postal(q.get('우편번호', ''))
            row['基本住所'] = clean_addr
            row['配送会社'] = '320'  # 사가와
            row['注文配送運賃タイプ'] = '10'  # 선불
            row['注文先名'] = q.get('수취인명', '')
            row['注文先電話番号'] = tel
            row['注文先国コード'] = 'JPN'
            row['注文先郵便番号'] = normalize_postal(q.get('우편번호', ''))
            row['注文先基本住所'] = clean_addr
            # 정렬용 내부 키 (마지막에 제거됨)
            row['_sort_cart'] = cart_no
            row['_sort_order'] = order_no
            row['_sort_sku'] = sku_code
            outbound_rows.append(row)

    # Power Query와 동일한 정렬: 장바구니 ASC → 주문 ASC → SKU ASC
    outbound_rows.sort(key=lambda r: (r['_sort_cart'], r['_sort_order'], r['_sort_sku']))
    for r in outbound_rows:
        r.pop('_sort_cart', None)
        r.pop('_sort_order', None)
        r.pop('_sort_sku', None)

    return outbound_rows, errors, addr_changes


def compute_audit(qsm_rows: List[Dict], outbound_rows: List[Dict],
                  mappings: Dict) -> Dict:
    """
    표1 시트 검수 지표 (row 2~5) 계산.
    OMS 업로드 결과와 수치 비교하는 용도.
    """
    # enabled QSM 행만
    enabled_qsm = []
    for q in qsm_rows:
        name = (q.get('상품명') or '').strip()
        option = (q.get('옵션정보') or '').strip()
        m = mappings.get((name, option))
        if m and m['enabled']:
            enabled_qsm.append((q, m))

    # 1. 총 상품 수량 = SUM of mapping 수량 합 (template D열 = 수량 컬럼)
    total_item_qty = 0
    # 2. 주문 업로드 개수 = SUM of SKU 개수 (output row 수)
    upload_row_count = 0
    for _, m in enabled_qsm:
        for sku_code, sku_qty in zip(m['sku_codes'], m['quantities']):
            if sku_code and sku_code != '-':
                total_item_qty += sku_qty
                upload_row_count += 1

    # 3. 송장번호 개수 = unique 장바구니번호 (enabled QSM)
    unique_carts = len({(q.get('장바구니번호') or '').strip() for q, _ in enabled_qsm
                        if (q.get('장바구니번호') or '').strip()})
    # 4. 주문번호 개수 = unique QSM 주문번호 (enabled QSM)
    unique_orders = len({(q.get('주문번호') or '').strip() for q, _ in enabled_qsm
                         if (q.get('주문번호') or '').strip()})

    # 실제 출고 PCS = SUM of 予定数量 (QSM수량 × 매핑수량)
    total_picking_pcs = sum(int(r.get('予定数量') or 0) for r in outbound_rows)

    # outbound_rows 검증
    outbound_carts = len({r['注文番号'] for r in outbound_rows if r.get('注文番号')})
    outbound_rows_count = len(outbound_rows)

    return {
        'total_item_qty': total_item_qty,              # 총 상품 수량 (매핑 수량 합)
        'upload_row_count': upload_row_count,          # 주문 업로드 개수 (KSE row)
        'unique_carts': unique_carts,                  # 송장번호 개수
        'unique_orders': unique_orders,                # 주문번호 개수 (QSM)
        'total_picking_pcs': total_picking_pcs,        # 실제 출고 PCS (×QSM수량)
        'check_total_match_count': total_item_qty == upload_row_count,
        'check_carts_match': unique_carts == outbound_carts,
        'check_rows_match': upload_row_count == outbound_rows_count,
    }


def build_outbound_xlsx(outbound_rows: List[Dict]) -> bytes:
    """
    원본 템플릿을 로드해 서식(컬럼 너비, 헤더 스타일, 폰트, 색상 등)을 그대로 보존한 채
    데이터 행만 교체하여 bytes 반환.
    """
    wb = openpyxl.load_workbook(OUTBOUND_TEMPLATE)
    ws = wb.active  # "Excel Sample"

    # 기존 샘플 데이터 행 삭제 (row 2 ~ max_row). 헤더(row1)는 유지.
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # 헤더(row1)의 각 컬럼 셀 서식을 template_style로 기억
    # → 데이터 행의 기본 서식으로 사용할 수 있음. 그러나 원본에선
    # 데이터 행이 별도 서식(맑은 고딕 등)이었으므로 그걸 재현하기 위해
    # 원본 row2 스타일 템플릿을 미리 보관.
    # 이미 delete_rows로 지웠으므로, 원본을 다시 읽어서 row2 스타일을 가져온다.
    style_wb = openpyxl.load_workbook(OUTBOUND_TEMPLATE)
    style_ws = style_wb.active
    data_styles = []
    for c in range(1, style_ws.max_column + 1):
        src = style_ws.cell(2, c)
        data_styles.append({
            'font': copy.copy(src.font),
            'fill': copy.copy(src.fill),
            'alignment': copy.copy(src.alignment),
            'border': copy.copy(src.border),
            'number_format': src.number_format,
        })
    style_wb.close()

    # 데이터 행 추가
    for ridx, row in enumerate(outbound_rows, start=2):
        for c, (jp, _) in enumerate(OUTBOUND_HEADERS, 1):
            cell = ws.cell(ridx, c, row.get(jp, ''))
            s = data_styles[c - 1] if c - 1 < len(data_styles) else None
            if s:
                cell.font = s['font']
                cell.fill = s['fill']
                cell.alignment = s['alignment']
                cell.border = s['border']
                cell.number_format = s['number_format']

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_qsm_waybill_csv(brief_content: bytes, waybill_map: Dict[str, str]) -> bytes:
    """
    brief.csv bytes + 장바구니번호→송장번호 매핑 → QSM 업로드용 CSV bytes.
    waybill_map: {장바구니번호: 송장번호}
    """
    text = brief_content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    fieldnames = reader.fieldnames
    out_rows = []
    missing = []
    for r in reader:
        cart_no = r.get('장바구니번호', '').strip()
        waybill = waybill_map.get(cart_no)
        if waybill:
            r['송장번호'] = waybill
        else:
            missing.append(cart_no)
        out_rows.append(r)

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(out_rows)
    return buf.getvalue().encode('utf-8-sig'), missing
